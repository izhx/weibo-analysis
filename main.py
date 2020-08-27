"""
微博情感分析，LDA主题聚类
"""

import os
import re
# import time
import pickle
import logging
from math import isnan
from codecs import open
from typing import List
from argparse import ArgumentParser
from datetime import datetime
from multiprocessing import Pool, Process, Queue

from gensim.corpora import Dictionary
from gensim.models import LdaModel, LdaMulticore
from pandas import DataFrame, Series, read_excel
from openpyxl import Workbook

LTP_DIR = "/data/private/zms/model_weight/ltp4-base/"

with open(f"{LTP_DIR}/vocab.txt") as file:
    CHAR = set(w.strip() for w in file.readlines())
    CHAR = set(w for w in CHAR if len(w) == 1)

with open("./stopwords.txt") as file:
    STOP_WORDS = set(w.strip() for w in file.readlines()) | CHAR

FORWARD_SPLIT = re.compile(r"//@[^/:：]+[:：]")
FORWARD_CONTENT = re.compile(r"//@[^/:：]+[:：][^/]+")
URL_REGEX = re.compile(r"http[s]?://[a-zA-Z0-9.?/&=:]*")

_ARG_PARSER = ArgumentParser(description="我的实验，需要指定配置文件")
_ARG_PARSER.add_argument('--name', '-n',
                         type=str,
                         default='clean',
                         help='configuration file path.')
_ARG_PARSER.add_argument('--visible', '-v',
                         type=str,
                         default='2,3',
                         help='gpu ids, like: 1,2,3')
_ARG_PARSER.add_argument('--range', '-r',
                         type=str,
                         default='5,21',
                         help='话题数搜索范围，左闭右开')
_ARG_PARSER.add_argument('--passes', '-p',
                         type=int,
                         default=50,
                         help='数据集迭代次数, epoch')
_ARG_PARSER.add_argument('--iterations', '-it',
                         type=int,
                         default=50,
                         help='推断时最大迭代次数')
_ARG_PARSER.add_argument('--keywords_num', '-k',
                         type=int,
                         default=50,
                         help='存储关键词数量')
_ARG_PARSER.add_argument('--pool_size', '-ps',
                         type=int,
                         default=16,
                         help='进程池大小')
_ARG_PARSER.add_argument('--debug', '-d', type=bool, default=False)
_ARGS = _ARG_PARSER.parse_args()

os.environ['OMP_NUM_THREADS'] = '1'
os.environ['TOKENIZERS_PARALLELISM'] = 'true'
os.environ['CUDA_VISIBLE_DEVICES'] = _ARGS.visible

# logging.basicConfig(format='[%(asctime)s - %(levelname)s] : %(message)s', level=logging.DEBUG)


def output(*args):
    message = ''.join([str(arg) for arg in args])
    print(f"[{str(datetime.now())[:-7]} - {os.getpid()}] {message}")


def generate_batch(data, batch_size):
    iterable = range(0, len(data) - batch_size, batch_size)
    if _ARGS.debug and os.getpid() == os.getppid():
        from tqdm import tqdm
        iterable = tqdm(iterable, total=len(data) // batch_size)
    for i in iterable:
        yield data[i: i + batch_size]
    else:
        yield data[i + batch_size: len(data)]


def ltp_tokenzie(texts, batch_size=128) -> List[List[str]]:
    import torch
    from ltp import LTP

    ltp = LTP(LTP_DIR, torch.device('cuda:1'))
    result = list()
    try:
        output(f"batch size {batch_size}, tokenize started...")
        for batch in generate_batch(texts, batch_size):
            # t = time.time()
            tokens, _ = ltp.seg(batch)  # batch 128, 1.8s
            result.extend(tokens)
            # output(time.time() - t)
        output("tokenize compete.")
    except RuntimeError as e:
        output('分词进程错误', e)
    return result


def _skep_feature(q, texts, batch_size):
    from paddlehub.reader.batching import pad_batch_data
    from paddlehub.reader.tokenization import convert_to_unicode, FullTokenizer

    def skep_batch(texts, tokenizer, max_seq_len=512):
        texts = [convert_to_unicode(t) for t in texts]
        tokens = [tokenizer.tokenize(t) for t in texts]
        # Account for [CLS] and [SEP] with "- 2"
        max_token = max_seq_len - 2
        tokens = [t[:max_token] if len(t) > max_token else t for t in tokens]
        # add [CLS] and [SEP], and convert to index
        tokens = [["[CLS]"] + t + ["[SEP]"] for t in tokens]
        token_ids = [tokenizer.convert_tokens_to_ids(t) for t in tokens]
        # get additional ids
        position_ids = [list(range(len(t))) for t in token_ids]
        text_type_ids = [[0] * len(t) for t in token_ids]
        task_ids = [[0] * len(t) for t in token_ids]

        pad_id = tokenizer.vocab["[PAD]"]
        padded_token_ids, input_mask = pad_batch_data(token_ids,
                                                      max_seq_len=max_seq_len,
                                                      pad_idx=pad_id,
                                                      return_input_mask=True)
        padded_text_type_ids = pad_batch_data(text_type_ids,
                                              max_seq_len=max_seq_len,
                                              pad_idx=pad_id)
        padded_position_ids = pad_batch_data(position_ids,
                                             max_seq_len=max_seq_len,
                                             pad_idx=pad_id)
        padded_task_ids = pad_batch_data(task_ids,
                                         max_seq_len=max_seq_len,
                                         pad_idx=pad_id)

        feature = [
            padded_token_ids, padded_position_ids, padded_text_type_ids,
            input_mask, padded_task_ids
        ]
        return feature

    tokenizer = FullTokenizer(
        "./ernie_skep_sentiment_analysis/assets/ernie_1.0_large_ch.vocab.txt")

    output("SKEP feature transform started.")
    for batch in generate_batch(texts, batch_size):
        feature = skep_batch(batch, tokenizer)
        q.put(feature)
    output("SKEP feature compete.")


def skep_analysis(texts, batch_size=16):
    def _score(result):
        return result['positive_probs'] - result['negative_probs']

    import numpy as np
    import paddlehub
    from paddle.fluid.core import PaddleTensor

    # 子进程处理特征
    q = Queue()
    p = Process(target=_skep_feature, args=(q, texts, batch_size))
    p.start()

    # 默认使用os.environ["CUDA_VISIBLE_DEVICES"]的第一个GPU
    skep = paddlehub.Module(directory="./ernie_skep_sentiment_analysis/")

    scores, i = np.zeros(len(texts)), 0
    output(f"batch size {batch_size}, sentiment analysis started...")
    while not q.empty():
        feature = q.get()
        inputs = [PaddleTensor(ndarray) for ndarray in feature]
        tensor = skep.predictor.run(inputs)[0]
        probs = np.array(tensor.data.float_data()).reshape(tensor.shape)
        scores[i: i + tensor.shape[0]] = probs[:, 1] - probs[:, 0]
        i += tensor.shape[0]
    p.join()
    output("sentiment analysis compete.")
    return scores


def read(path) -> DataFrame:
    def _clean(row):
        text = URL_REGEX.sub('', row.contents)
        if row.is_forward and '//@' in text:
            # 如果是转发的且格式正确
            if text.startswith('//@'):
                # 如果单纯转发，则内容设置为最原始微博的内容
                try:
                    text = FORWARD_CONTENT.findall(text)[-1]
                    i = FORWARD_SPLIT.match(text).regs[0][1]
                    text = text[i:]
                except IndexError:
                    text = text.replace('//@', '')  # TODO 可以用weibo的API处理
            else:
                # 否则截取新内容
                text = text[:text.find('//@')]
        return text

    output(f"===> Reading from <{path}>.")
    data: DataFrame = read_excel(path)  # .iloc[:280]

    # 只保留想要的4列，并去除空值，截取日期
    data = data[['contents', 'time', 'id', 'is_forward']].dropna().reset_index()
    data['date'] = data['time'].apply(lambda s: s[:10])
    data['contents'] = data['contents'].astype(str)

    # 预处理文本，并行分词和情感分析
    texts = data.apply(_clean, axis=1).to_list()
    pool = Pool(1)
    tokens = pool.apply_async(ltp_tokenzie, (texts, 192))
    data['sentiment_score'] = skep_analysis(texts, 16)
    data['tokens'] = tokens.get()
    pool.close()
    pool.join()
    return data[['date', 'tokens', 'id', 'sentiment_score']]


def to_excel(topics, data: DataFrame, topic_ids):
    data['topic_id'] = topic_ids
    # 单独输出统计结果
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主题与关键词"
    for topic_id, (tokens, score) in enumerate(topics):
        c = topic_id * 3
        sheet.cell(1, c + 1).value = "id与分数:"
        sheet.cell(1, c + 2).value = topic_id
        sheet.cell(1, c + 3).value = float(score)
        sheet.cell(2, c + 1).value = "关键词"
        sheet.cell(2, c + 2).value = "词频"
        sheet.cell(2, c + 3).value = "权重"
        for r, (token, weight, freq) in enumerate(tokens, 3):
            sheet.cell(r, c + 1).value = token
            sheet.cell(r, c + 2).value = freq
            sheet.cell(r, c + 3).value = float(weight)

    dates = sorted(data['date'].unique())
    head = ['日期', '总'] + [f"topic_{i}" for i in range(len(topics))]

    num_sheet = workbook.create_sheet('微博数')
    num_sheet.append(head)
    sent_sheet = workbook.create_sheet('情感分')
    sent_sheet.append(head)

    for r, date in enumerate(dates, 2):
        num_sheet.cell(r, 1).value = date
        sent_sheet.cell(r, 1).value = date
        today = data[data.date == date]
        num_sheet.cell(r, 2).value = len(today)
        score = today['sentiment_score'].mean()
        sent_sheet.cell(r, 2).value = score
        for tid in range(len(topics)):
            today_topic = today.query(f"topic_id == {tid}")
            num_sheet.cell(r, tid + 3).value = len(today_topic)
            today_topic_score = today_topic['sentiment_score'].mean()
            if not isnan(today_topic_score):
                sent_sheet.cell(r, tid + 3).value = today_topic_score
    else:  # 写入总的
        r += 1
        num_sheet.cell(r, 1).value = '平均'
        sent_sheet.cell(r, 1).value = '平均'
        num_sheet.cell(r, 2).value = len(data)
        score = data['sentiment_score'].mean()
        sent_sheet.cell(r, 2).value = score
        for tid in range(len(topics)):
            _topic = data.query(f"topic_id == {tid}")
            num_sheet.cell(r, tid + 3).value = len(_topic)
            _topic_score = _topic['sentiment_score'].mean()
            sent_sheet.cell(r, tid + 3).value = _topic_score

    sheet = workbook.create_sheet('id与topic_id')
    data.apply(lambda x: sheet.append([x.id, int(x.topic_id)]), axis=1)

    path = f"./dev/excel/{_ARGS.name}_{len(topics)}.xlsx"
    workbook.save(path)
    output(f"result saved at <{path}>")


def save_and_inference(model: LdaModel, corpus, num_topics):
    path = f"./dev/model/covid_{_ARGS.name}_{num_topics}.pkl"
    try:
        model.save(path)
        output(f"model saved at <{path}>")
        gamma, _ = model.inference(corpus)
    except RuntimeError as e:
        logging.error(f"PID: {os.getpid()}, num_topics: {num_topics} error")
        print(e)
    output(f"num_topics {num_topics} inference compete.")
    return gamma.argmax(axis=1)


def get_model(corpus, num_topics, kwargs):
    output(f"running num_topics: {num_topics}.")
    try:
        model = LdaModel(corpus, num_topics, **kwargs)
        topic_ids = save_and_inference(model, corpus, num_topics)
    except RuntimeError as e:
        logging.error(f"PID: {os.getpid()}, num_topics: {num_topics} error")
        print(e)
    return model, topic_ids


def pipline(data: DataFrame, cache: str):
    documents = data['tokens'].to_list()
    # Create a dictionary representation of the documents.
    dictionary = Dictionary(documents)

    # Filter out words that occur less than 20 documents, or more than 50% of the documents.
    dictionary.filter_extremes(no_below=20, no_above=0.5)

    # 去停用词
    bad_ids = [dictionary.token2id[t] for t in STOP_WORDS if t in dictionary.token2id]
    dictionary.filter_tokens(bad_ids=bad_ids)

    # Bag-of-words representation of the documents.
    corpus = [dictionary.doc2bow(doc) for doc in documents]

    _ = dictionary[0]  # This is only to "load" the dictionary.
    print('Number of unique tokens: ', len(dictionary))
    print('Number of documents: ', len(corpus))
    # test = get_model(6, corpus, dictionary.id2token)

    topic_range = tuple(int(s.strip()) for s in _ARGS.range.split(','))
    kwargs = dict(
        id2word=dictionary.id2token, chunksize=len(corpus),
        passes=_ARGS.passes, alpha='auto', eta='auto', eval_every=None,
        iterations=_ARGS.iterations, random_state=123)
    result_dict = dict()
    if len(corpus) < 5e5:  # 并行训练模型
        pool = Pool(_ARGS.pool_size)
        for k in range(*topic_range):
            result_dict[k] = pool.apply_async(get_model, (corpus, k, kwargs))
        result_dict = {k: v.get() for k, v in result_dict.items()}
    else:
        pool = Pool(1)
        for k in range(*topic_range):
            model = LdaMulticore(corpus, k, workers=_ARGS.pool_size, **kwargs)
            ids = pool.apply_async(save_and_inference, (model, corpus, k))
            result_dict[k] = [model, ids]
        result_dict = {k: (v[0], v[1].get()) for k, v in result_dict.items()}
    pool.close()  # 等子进程执行完毕后关闭进程池
    pool.join()

    output(f"Searched range{topic_range}")

    # 计算一致性的代码自己有多进程，所以只能串行
    for k, (model, ids) in result_dict.items():
        print('\nnum_topics: ', k)
        # print('Model perplexity: ', model.log_perplexity(corpus))  # 这个没做归一
        top_topics = model.top_topics(corpus, documents, dictionary,
                                      coherence='c_v', topn=_ARGS.keywords_num,
                                      processes=_ARGS.pool_size)
        scores = Series([t[1] for t in top_topics])
        print('Coherence Score: ', scores.mean())
        # 得到关键词词频
        topics_info = list()
        for _topic in top_topics:
            tokens = [(t[1], t[0], dictionary.cfs[dictionary.token2id[t[1]]]) for t in _topic[0]]
            topics_info.append((tokens, _topic[1]))
        to_excel(topics_info, data, ids)

    output(f"===> {_ARGS.name} compete. \n")


def main():
    def dump_cache(data, path):
        with open(path, mode='wb') as file:
            pickle.dump(data, file)
        output(f"data saved at <{path}>")

    def load_cache(path):
        with open(path, mode='rb') as file:
            data = pickle.load(file)
        output(f"data loaded from <{path}>")
        return data

    cache = f"./dev/data/{_ARGS.name}.pkl"
    if os.path.isfile(cache):
        df = load_cache(cache)
    else:
        if _ARGS.name == 'clean':
            df: DataFrame = None
            for i in range(6):
                path = f"./dev/data/covid19/clean{i}_covid19.xlsx"
                part = read(path)
                dump_cache(part, path.replace('.xlsx', '.pkl'))
                df = df.append(part) if df else part
        else:
            path = f"./dev/data/covid19/{_ARGS.name}_covid19.xlsx"
            df = read(path)

        dump_cache(df, cache)
        logging.disable(level=logging.INFO)

    pipline(df, cache)
    return


if __name__ == "__main__":
    main()

"""
统计每天每主题数量，每天（每主题）感情变化，
"""
