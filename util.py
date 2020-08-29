"""
公用代码
"""

import os
import pickle
from math import isnan
from typing import Tuple, Sequence

from pandas import DataFrame
from datetime import datetime
from openpyxl import Workbook

LTP_DIR = "/data/private/zms/model_weight/ltp4-base/"
SKEP_DIR = "./ernie_skep_sentiment_analysis/"
CACHE_DIR = "./dev/cache/"


def output(*args):
    message = ''.join([str(arg) for arg in args])
    print(f"[{str(datetime.now())[:-7]} - {os.getpid()}] {message}")


def cache_path(name):
    return os.path.join(CACHE_DIR, name + ".pkl")


def dump_cache(data, name):
    with open(cache_path(name), mode='wb') as file:
        pickle.dump(data, file)
    output(f"file saved at <{cache_path(name)}>")


def load_cache(name):
    with open(cache_path(name), mode='rb') as file:
        data = pickle.load(file)
    output(f"file loaded from <{cache_path(name)}>")
    return data


def generate_batch(data: Sequence, batch_size: int) -> Tuple[slice, Sequence]:
    indices = range(0, len(data) - batch_size, batch_size)
    if batch_size >= len(data):
        yield slice(0, len(data)), data
    for i in indices:
        _slice = slice(i, i + batch_size)
        yield _slice, data[_slice]
    _slice = slice(i + batch_size, len(data))
    yield _slice, data[_slice]


def to_excel(topics, data: DataFrame, topic_ids, name):
    """
    统计信息并写入
    """
    data['topic_id'] = topic_ids
    # 单独输出统计结果
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主题与关键词"
    for topic_id, (tokens, score) in enumerate(topics):
        c = topic_id * 3
        sheet.cell(1, c + 1).value = "id与分数:"
        sheet.cell(1, c + 2).value = topic_id
        sheet.cell(1, c + 3).value = float(score)
        sheet.cell(2, c + 1).value = "关键词"
        sheet.cell(2, c + 2).value = "词频"
        sheet.cell(2, c + 3).value = "权重"
        for r, (token, weight, freq) in enumerate(tokens, 3):
            sheet.cell(r, c + 1).value = token
            sheet.cell(r, c + 2).value = freq
            sheet.cell(r, c + 3).value = float(weight)

    dates = sorted(data['date'].unique())
    head = ['日期', '总'] + [f"topic_{i}" for i in range(len(topics))]

    num_sheet = workbook.create_sheet('微博数')
    num_sheet.append(head)
    sent_sheet = workbook.create_sheet('情感分')
    sent_sheet.append(head)

    for r, date in enumerate(dates, 2):
        num_sheet.cell(r, 1).value = date
        sent_sheet.cell(r, 1).value = date
        today = data[data.date == date]
        num_sheet.cell(r, 2).value = len(today)
        score = today['sentiment_score'].mean()
        sent_sheet.cell(r, 2).value = score
        for tid in range(len(topics)):
            today_topic = today.query(f"topic_id == {tid}")
            num_sheet.cell(r, tid + 3).value = len(today_topic)
            today_topic_score = today_topic['sentiment_score'].mean()
            if not isnan(today_topic_score):
                sent_sheet.cell(r, tid + 3).value = today_topic_score
    else:  # 写入总的
        r += 1
        num_sheet.cell(r, 1).value = '平均'
        sent_sheet.cell(r, 1).value = '平均'
        num_sheet.cell(r, 2).value = len(data)
        score = data['sentiment_score'].mean()
        sent_sheet.cell(r, 2).value = score
        for tid in range(len(topics)):
            _topic = data.query(f"topic_id == {tid}")
            num_sheet.cell(r, tid + 3).value = len(_topic)
            _topic_score = _topic['sentiment_score'].mean()
            sent_sheet.cell(r, tid + 3).value = _topic_score

    sheet = workbook.create_sheet('id与topic_id')
    data.apply(lambda x: sheet.append([x.id, int(x.topic_id)]), axis=1)

    path = f"./dev/excel/{name}_{len(topics)}.xlsx"
    workbook.save(path)
    print(f"==> result saved at <{path}>")
